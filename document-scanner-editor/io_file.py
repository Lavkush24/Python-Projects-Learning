#!/usr/bin/env python3
"""
Excel File I/O Handler
Handles reading from and writing to Excel files for course data validation
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Tuple, Optional, Union
import os
from datetime import datetime

class ExcelFileHandler:
    def __init__(self):
        # Define the expected column headers based on the data format
        self.expected_columns = [
            'Institution Id',
            'Course Id', 
            'Course Type',
            'Degree Type',
            'Course Name',
            'L3 Tagging',
            'Course Intake Ids',
            'Course start date',
            'Course Level Tuition Fees',
            'Application Fees',
            'Application Deadline Date',
            'Course Level URL',
            'Course Status',
            'Show',
            'Study Mode',
            'Previous Education',
            'Commissionable'
        ]
        
        # Define color schemes for different error types
        self.error_colors = {
            'Numeric': PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid'),  # Light red
            'Unique': PatternFill(start_color='FFB366', end_color='FFB366', fill_type='solid'),   # Orange
            'Capitalization': PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'),  # Light yellow
            'Blank': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'),    # Pink
            'Date': PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid'),     # Light blue
            'URL': PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'),      # Red
            'Count': PatternFill(start_color='CC99FF', end_color='CC99FF', fill_type='solid'),    # Purple
            'Status': PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')    # Blue
        }
        
        # Define styles
        self.header_font = Font(bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def read_excel_file(self, file_path: str) -> Tuple[pd.DataFrame, List[str]]:
        """
        Read data from Excel file
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Tuple of (DataFrame, List of column names)
        """
        try:
            # Read Excel file
            df = pd.read_excel(file_path, engine='openpyxl')
            
            # Get column names
            columns = df.columns.tolist()
            
            print(f"Successfully read Excel file: {file_path}")
            print(f"Found {len(df)} rows and {len(columns)} columns")
            print(f"Columns: {columns}")
            
            return df, columns
            
        except FileNotFoundError:
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")
    
    def read_csv_file(self, file_path: str) -> Tuple[pd.DataFrame, List[str]]:
        """
        Read data from CSV file
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            Tuple of (DataFrame, List of column names)
        """
        try:
            # Read CSV file
            df = pd.read_csv(file_path)
            
            # Get column names
            columns = df.columns.tolist()
            
            print(f"Successfully read CSV file: {file_path}")
            print(f"Found {len(df)} rows and {len(columns)} columns")
            print(f"Columns: {columns}")
            
            return df, columns
            
        except FileNotFoundError:
            raise FileNotFoundError(f"CSV file not found: {file_path}")
        except Exception as e:
            raise Exception(f"Error reading CSV file: {str(e)}")
    
    def normalize_column_names(self, columns: List[str]) -> Dict[str, str]:
        """
        Normalize column names to match expected format
        
        Args:
            columns: List of actual column names from the file
            
        Returns:
            Dictionary mapping actual column names to normalized names
        """
        column_mapping = {}
        
        # Common variations and abbreviations
        variations = {
            'Instituti': 'Institution Id',
            'Course I': 'Course Id',
            'Course TDegree': 'Degree Type',
            'Course': 'Course Name',
            'L3 Taggi': 'L3 Tagging',
            'Course sCourse': 'Course start date',
            'LApplicati': 'Course Level Tuition Fees',
            'Applicati': 'Application Fees',
            'Course LCourse SShow': 'Course Level URL',
            'Study M': 'Study Mode',
            'Previous': 'Previous Education',
            'Commissionable': 'Commissionable'
        }
        
        for col in columns:
            # Try exact match first
            if col in self.expected_columns:
                column_mapping[col] = col
                continue
            
            # Try variations
            normalized = None
            for variation, expected in variations.items():
                if col.startswith(variation) or variation in col:
                    normalized = expected
                    break
            
            if normalized:
                column_mapping[col] = normalized
            else:
                # Keep original name if no match found
                column_mapping[col] = col
        
        return column_mapping
    
    def clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean and preprocess the data
        
        Args:
            df: Raw DataFrame
            
        Returns:
            Cleaned DataFrame
        """
        # Create a copy to avoid modifying original
        cleaned_df = df.copy()
        
        # Handle NULL values
        cleaned_df = cleaned_df.replace(['NULL', 'NUL', 'null', 'nul'], '')
        
        # Clean up date formats (remove extra numbers after date)
        if 'Course start date' in cleaned_df.columns:
            cleaned_df['Course start date'] = cleaned_df['Course start date'].astype(str).apply(
                lambda x: self._clean_date_format(x)
            )
        
        # Clean up Study Mode
        if 'Study Mode' in cleaned_df.columns:
            cleaned_df['Study Mode'] = cleaned_df['Study Mode'].astype(str).apply(
                lambda x: self._normalize_study_mode(x)
            )
        
        # Clean up Course Status and Show (extract from combined field)
        if 'Course Level URL' in cleaned_df.columns:
            cleaned_df = self._extract_status_and_show(cleaned_df)
        
        return cleaned_df
    
    def _clean_date_format(self, date_str: str) -> str:
        """Clean date format by removing extra numbers"""
        if pd.isna(date_str) or date_str == '':
            return ''
        
        # Remove extra numbers after the date (e.g., "2024-09-015720" -> "2024-09-01")
        date_str = str(date_str).strip()
        
        # Try to extract YYYY-MM-DD format
        import re
        date_match = re.search(r'(\d{4}-\d{2}-\d{2})', date_str)
        if date_match:
            return date_match.group(1)
        
        return date_str
    
    def _normalize_study_mode(self, mode: str) -> str:
        """Normalize study mode values"""
        if pd.isna(mode) or mode == '':
            return ''
        
        mode = str(mode).strip()
        
        if 'full' in mode.lower():
            return 'Full time'
        elif 'part' in mode.lower():
            return 'Part time'
        
        return mode
    
    def _extract_status_and_show(self, df: pd.DataFrame) -> pd.DataFrame:
        """Extract Course Status and Show from combined field"""
        if 'Course Level URL' not in df.columns:
            return df
        
        # Create new columns if they don't exist
        if 'Course Status' not in df.columns:
            df['Course Status'] = ''
        if 'Show' not in df.columns:
            df['Show'] = ''
        
        # Extract from combined field (assuming format like "https://wvOpen, Ope No")
        for idx, row in df.iterrows():
            combined_value = str(row['Course Level URL'])
            
            # Extract URL part (before any comma)
            if ',' in combined_value:
                parts = combined_value.split(',')
                url_part = parts[0].strip()
                status_part = ','.join(parts[1:]).strip()
                
                # Update URL
                df.at[idx, 'Course Level URL'] = url_part
                
                # Extract status and show
                if 'Open' in status_part:
                    df.at[idx, 'Course Status'] = 'Open'
                elif 'Closed' in status_part:
                    df.at[idx, 'Course Status'] = 'Closed'
                
                if 'Yes' in status_part:
                    df.at[idx, 'Show'] = 'Yes'
                elif 'No' in status_part:
                    df.at[idx, 'Show'] = 'No'
        
        return df
    
    def write_excel_with_validation_results(self, 
                                          df: pd.DataFrame, 
                                          validation_errors: List[Dict], 
                                          output_path: str = None) -> str:
        """
        Write DataFrame to Excel with validation error highlighting
        
        Args:
            df: DataFrame to write
            validation_errors: List of validation error dictionaries
            output_path: Output file path (optional)
            
        Returns:
            Path to the created Excel file
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"validated_course_data_{timestamp}.xlsx"
        
        try:
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Validated Data"
            
            # Write headers
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.border
                    
                    # Apply error highlighting
                    self._apply_error_highlighting(cell, row_idx, col_idx, validation_errors)
            
            # Add error count column
            error_count_col = len(df.columns) + 1
            ws.cell(row=1, column=error_count_col, value="Error Count").font = self.header_font
            ws.cell(row=1, column=error_count_col).fill = self.header_fill
            ws.cell(row=1, column=error_count_col).border = self.border
            
            # Calculate error counts per row
            for row_idx in range(2, len(df) + 2):
                row_errors = [e for e in validation_errors if e['row'] == row_idx - 1]
                error_count = len(row_errors)
                
                cell = ws.cell(row=row_idx, column=error_count_col, value=error_count)
                cell.border = self.border
                
                if error_count > 0:
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    cell.font = Font(color='FFFFFF', bold=True)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
            # Create summary sheet
            self._create_summary_sheet(wb, df, validation_errors)
            
            # Create error details sheet
            self._create_error_details_sheet(wb, validation_errors)
            
            # Save the workbook
            wb.save(output_path)
            print(f"Excel file saved successfully: {output_path}")
            
            return output_path
            
        except Exception as e:
            raise Exception(f"Error writing Excel file: {str(e)}")
    
    def _apply_error_highlighting(self, cell, row_idx: int, col_idx: int, validation_errors: List[Dict]):
        """Apply error highlighting to cells"""
        # Find errors for this specific cell
        for error in validation_errors:
            if error['row'] == row_idx - 1:  # Convert to 0-based index
                # Map column index to column name (this is a simplified approach)
                # In a real implementation, you'd need to map column indices to names
                if error['error_type'] in self.error_colors:
                    cell.fill = self.error_colors[error['error_type']]
                    cell.font = Font(bold=True)
                    break
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            print(f"Warning: Could not auto-adjust columns: {e}")
    
    def _create_summary_sheet(self, wb, df: pd.DataFrame, validation_errors: List[Dict]):
        """Create summary sheet with validation statistics"""
        ws = wb.create_sheet("Validation Summary")
        
        # Summary statistics
        ws['A1'] = "Validation Summary"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = "Total Rows:"
        ws['B3'] = len(df)
        
        ws['A4'] = "Total Errors:"
        ws['B4'] = len(validation_errors)
        
        ws['A5'] = "Rows with Errors:"
        rows_with_errors = len(set(error['row'] for error in validation_errors))
        ws['B5'] = rows_with_errors
        
        # Error breakdown by type
        ws['A7'] = "Error Breakdown by Type:"
        ws['A7'].font = Font(bold=True)
        
        error_types = {}
        for error in validation_errors:
            error_type = error['error_type']
            error_types[error_type] = error_types.get(error_type, 0) + 1
        
        row = 8
        for error_type, count in error_types.items():
            ws[f'A{row}'] = error_type
            ws[f'B{row}'] = count
            row += 1
    
    def _create_error_details_sheet(self, wb, validation_errors: List[Dict]):
        """Create error details sheet"""
        ws = wb.create_sheet("Error Details")
        
        # Headers
        headers = ['Row', 'Column', 'Error Type', 'Message', 'Value']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Error details
        for row_idx, error in enumerate(validation_errors, 2):
            ws.cell(row=row_idx, column=1, value=error['row'])
            ws.cell(row=row_idx, column=2, value=error['column'])
            ws.cell(row=row_idx, column=3, value=error['error_type'])
            ws.cell(row=row_idx, column=4, value=error['message'])
            ws.cell(row=row_idx, column=5, value=error['value'])
            
            # Apply border to all cells in the row
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).border = self.border
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)

# Create a global instance for easy use
excel_handler = ExcelFileHandler()
